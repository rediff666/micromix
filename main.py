import os
import ffmpeg
import audioread
from openpyxl import Workbook, load_workbook 

from hojadrive import worksheet
from voztrack import VozTrack

# Ejemplo de uso
input_file = '4444.mp3'
input_file="C:/Users/haro/Desktop/System HO/softjh/MicroMixes_Fusion/Salida/kk.mp3"
output_file = 'Salida/kk2.mp3'
start_time = '00:00:30'  # Extraer desde el segundo 30
duracion = '00:00:10'  # Durante 10 segundos



# # Ejemplo de uso
input_file1 = '4444.mp3'
input_file2 = '2222.mp3'
# output_file = 'output.mp3'

def get_duracion_audio(audio_in):
	with audioread.audio_open(audio_in) as audio_cargado:
		duracion=audio_cargado.duration
	return duracion

def sampleador(audio_in,inicio,duracion,audio_final):
	ffmpeg.input(audio_in,ss=inicio,t=duracion).output(audio_final).run(overwrite_output=True)


def mixear_audios(audios_in,audio_final):

	audios_in_pre=[]
	for audio_in in audios_in:
		audios_in_pre.append(ffmpeg.input(audio_in))

	ffmpeg.concat(*audios_in_pre,v=0,a=1).output(audio_final).run(overwrite_output=True)

def ejecutor_total():
	partes=input("En cuantas partes se dividiran los audios")
	duracion=input("Que duración ")


def get_reloj(tiempo):

	if tiempo<60:
		if tiempo>10: segundos=tiempo 
		else: segundos=f"0{str(tiempo)}"
		reloj=f"00:00:{segundos}"
	else:
		hora="00" if tiempo<3600 else str(int(tiempo/3600))
		hora=f"0{hora}" if int(hora)<10 else hora

		minutos_pre=int(tiempo%3600)
		minutos_pre=int(minutos_pre/60) #if tiempo<3600 else int(tiempo%3600)

		minutos=f"0{minutos_pre}" if minutos_pre<10 else minutos_pre

		segundos_pre=tiempo%60
		if segundos_pre>10: segundos=segundos_pre
		else: segundos=f"0{segundos_pre}"

		reloj=f"{hora}:{minutos}:{segundos}"
	return reloj



def get_momentos(partes,duracion,audio_in,offset=3,tipo_intervalos=1):
	duracion_segundos=get_duracion_audio(audio_in)
	intervalo_tiempo=int(int(duracion_segundos)/partes)
	momentos=[]
	array_intervalos=[]
	if tipo_intervalos==1:
		for _ in range(0,int(partes)):
			momento=_*intervalo_tiempo
			array_intervalos.append(momento)
	else:
		for _ in range(0,int(partes)):
			# if _==0:momento=15
			# else: momento=+15
			momento=25 if _==0 else momento+25
			array_intervalos.append(momento)

	for _ in enumerate(array_intervalos):
		momento=_[1]
		if _[0]==0:
			# if tipo_intervalos==1:
			momento=momento+offset # Para el primer sample

		if momento<60:
			minutos="00"
			if momento>=10: segundos=momento 
			else: segundos=f"0{momento}"
		else:
			minutos_pre=int(round(momento/60,1))
			if minutos_pre>=10:	minutos=f"{minutos_pre}" 
			else:minutos=f"0{minutos_pre}"

			segundos_pre=momento%60
			if segundos_pre>=10: segundos=segundos_pre
			else: segundos=f"0{segundos_pre}"
		
		mm=f"00:{minutos}:{segundos}"
		momentos.append(mm)

	
	return momentos

def crea_mix(partes,duracion):
	libro=load_workbook("base_micromix.xlsx",keep_vba=False)
	hoja=libro['Hoja1']
	n_fila=0
	array_submixes=[]
	
	archivo_srt=open("subtitulos.srt","w",encoding="utf-8")
	archivo_tiempo= open("reloj.txt","w",encoding="utf-8")
	archivo_log=open("logs.txt","w",encoding="utf-8")

	archivo_log.write(f"Archivos que generaron error\n")

	### Varlores para Drive
	matriz_valores=[]

	### Offset, Para el primer sample, cuando iniciará
	offset=int(input("Cuantos segundos adicionales offset para extraer primer sample en track total:\n"))
	off_ms=int(input("En cuantos segundos inicia la voz de Track (Defecto:0): "))
	tipo_intervalos=int(input("Elija Tipo de intevalo de samples de submixes:\n1 Division de Partes\n2 Tramos Fijos\n"))

	for fila in hoja.iter_rows(min_row=2,max_col=10,max_row=1000):
		n_fila+=1
		fila_array=[]	

		if str(fila[0].value)!="None":
			# print(f"{fila[0].value} --- {fila[1].value}")
			audio_file=f"{fila[0].value}/{fila[1].value}"

			try:

				momentos=get_momentos(partes,duracion,audio_file,offset,tipo_intervalos)
				audio_final="sample"
				array_samples=[]

				for _ in range(0,len(momentos)):
					array_samples.append(f"Submixes/{audio_final}{_+1}.mp3")
					sampleador(audio_file,momentos[_],duracion,f"Submixes/{audio_final}{_+1}.mp3")

					### Creación de voz para sample, Para primer Fragmento de Sample 
					# if _==0:
					# 	# vv=VozTrack(fila[1].value)
					# 	# vv.crea_sample2()
					# 	pass

				mixear_audios(array_samples,f"Submixes/submix{n_fila}.mp3")
				array_submixes.append(f"Submixes/submix{n_fila}.mp3")

				###	Creación de voz para sample, Para todo el Submix
				vv=VozTrack(fila[1].value)
				vv.crea_sample(f"Submixes/submix{n_fila}.mp3",off_ms)

				instante=get_reloj((n_fila-1)*partes*duracion)
				instante2_subtitulo=get_reloj((n_fila)*(partes*duracion))

				### Escrituras a Archivos y a Drive
				archivo_srt.write(f"{n_fila}\n")
				archivo_srt.write(f"{instante},000 --> {instante2_subtitulo},000\n")
				archivo_srt.write(f"{fila[1].value}\n\n")

				archivo_tiempo.write(f"{instante} -> {fila[0].value} --- {fila[1].value} \n")

				fila_array=[n_fila,instante,fila[1].value]
				matriz_valores.append(fila_array)

				for _ in array_samples:
					os.remove(_)
			except:
				archivo_log.write(f"* {n_fila} --> {audio_file}\n")
			
		else:
			break

	
	###  Mix Final
	mixear_audios(array_submixes,f"Mix/mixxxx.mp3")

	### Borrado de submixes
	for _ in array_submixes:
		os.remove(_)

	### Valores a Base en Drive
	drive=worksheet()
	drive.insertar_filas_en_base(matriz_valores)
	matriz_valores.clear()

	### Cierre de archivo
	archivo_srt.close()
	archivo_tiempo.close()
	archivo_log.close()

def main():
	print("*********** CREADOR DE MIXES ***************")
	partes=input("Cuantos samples por track?? \n")
	try: partes=int(partes) 
	except: partes=2
	duracion=input("Cuanto dura cada sample??? \n")
	try: duracion=int(duracion) 
	except: duracion=7

	crea_mix(partes,duracion)

main()



# def pruebas_de_tiempos(audio_in):
# 	tiempo_decimales=get_duracion_audio(audio_in)
# 	print(tiempo_decimales)
# 	divisor=round(int(tiempo_decimales)/60,2)
# 	if divisor<60:
# 		print("No dura mas de una hora")
# 	minutos=str(divisor).split(".")[0]
# 	segundos=(int(str(divisor).split(".")[-1])/100)*60
# 	print(f"minutos {minutos}")
# 	print(f"segundos {segundos}")


# # get_momentos(2)
# # sampleador(input_file,start_time,duracion,output_file)



# # mixear_audios([input_file1,input_file2],"salida.mp3")